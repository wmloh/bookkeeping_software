# Instructions at http://openpyxl.readthedocs.io/en/default/usage.html

import sys
import pygame
import time
import math
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet import *
from openpyxl.cell import Cell
from openpyxl.styles import *
from openpyxl.descriptors.serialisable import Serialisable
import csv
from tkinter import filedialog
import datetime
from datetime import date

line = 0

toggle = 0
selector = True

all_asset = ['Cash','Account Receivable','Notes Receivable','Buildings','Land','Equipment',
 'Prepaid Expenses','Supplies','Inventory']
all_liab = ['Account Payable','Notes Payable','Interest Payable','Taxes Payable',
 'Bonds Payable','Deferred Revenue','Loan Payable','Accrued Liabilities']
all_SHE = ['Retained Earnings','Common Shares','Dividends','Contributed Capital']
all_expenses = ['Cost of Goods Sold','Salary Expense', 'Utilities Expense', 'Rent Expense',
            'Insurance Expense']
all_revenue = ['Sales Revenue', 'Service Revenue']

asset = ['Cash','Account Receivable','Notes Receivable','Buildings','Land','Equipment',
 'Prepaid Expenses','Supplies','Inventory']
liab = ['Account Payable','Notes Payable','Interest Payable','Taxes Payable',
 'Bonds Payable','Deferred Revenue','Loan Payable','Accrued Liabilities']
SHE = ['Retained Earnings','Common Shares','Dividends','Contributed Capital']
expenses = ['Cost of Goods Sold','Salary Expense', 'Utilities Expense', 'Rent Expense',
            'Insurance Expense']
revenue = ['Sales Revenue', 'Service Revenue']

other_asset = []
other_liab = []
other_SHE = []
other_expenses = []
other_revenue = []

accounts_set = set()

dest_filename = 'Test2'

black = (0,0,0)
white = (255,255,255)
gray = (155,155,155)
blue = (0,35,170)
light_blue = (10,190,255)
green = (0,215,15)
dark_green = (0,120,5)
yellow = (255,255,0)
dark_yellow = (195,195,0)
red = (255,0,0)
dark_red = (120,0,0)

pygame.init()

display_width = 1366
display_height = 768
gameDisplay = pygame.display.set_mode((display_width,display_height), pygame.FULLSCREEN)
pygame.display.set_caption('Bookkeeping Automator')
#pygame.display.set_icon(icon)
clock = pygame.time.Clock()

transparency = pygame.image.load("Transparency.png")

tinyfont = pygame.font.SysFont('georgia', 15)
smallfont = pygame.font.SysFont('georgia', 16)
medfont = pygame.font.SysFont('georgia', 18)
largefont = pygame.font.SysFont('georgia', 26)

element_d = ''
account_d = ''
element_c = ''
account_c = ''
value = ''
date = ''
description = ''
tab = 0
comp_name = ''
fs_date = ''
transaction_list = []
dep_accounts = dict()
depreciation_account = ''
dep_method = ''
dep_ac_cost = 0
dep_rs_cost = 0
dep_time = 0
dep_unit = 0

def text_objects(text, color,size):
    if size == 'small':
        textSurface = smallfont.render(text, True, color)
    elif size == 'medium':
        textSurface = medfont.render(text, True, color)
    elif size == 'large':
        textSurface = largefont.render(text, True, color)
    elif size == 'tiny':
        textSurface = tinyfont.render(text, True, color)
    return textSurface, textSurface.get_rect()

def text_to_button(text, color, buttonx, buttony, buttonwidth, buttonheight,size = 'small', type = 'quad'):
    if type == 'quad':
        textSurf, textRect = text_objects(text, color, size)
        textRect.center = ((buttonx+(buttonwidth/2)),buttony + (buttonheight/2))
        gameDisplay.blit(textSurf,textRect)
    elif type == 'circle':
        textSurf, textRect = text_objects(text, color, size)
        textRect.center = ((buttonx, buttony))
        gameDisplay.blit(textSurf, textRect)

def find_row(column, row, ws, space = 0):
    while not ws[column + str(row)].value == None:
        row += 1
    if space == 0:
        return row
    else:
        return find_row(column, row+1, ws, space = space - 1)

def debit_stage():
    global element_d
    global account_d
    global account_c
    global element_c
    global toggle
    global date
    global value
    global description
    global tab
    debit = True
    toggle = 0
    while debit:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                quit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    pygame.quit()
                    quit()
                elif event.key == pygame.K_SPACE:
                    toggle += 1
                    account_d = ''
                elif event.key == pygame.K_1:
                    if toggle % 2 == 0:
                        element_d = 'Assets'
                    elif element_d == 'Assets':
                        account_d = 'Cash'
                    elif element_d == 'Liabilities':
                        account_d = 'Account Payable'
                    elif element_d == "Shareholders' Equity":
                        account_d = 'Retained Earnings'
                    elif element_d == "Expenses":
                        account_d = 'Cost of Goods Sold'
                    elif element_d == "Revenue":
                        account_d = revenue[0]
                elif event.key == pygame.K_2:
                    if toggle % 2 == 0:
                        element_d = 'Liabilities'
                    elif element_d == 'Assets':
                        account_d = 'Account Receivable'
                    elif element_d == 'Liabilities':
                        account_d = 'Notes Payable'
                    elif element_d == "Shareholders' Equity":
                        account_d = 'Common Shares'
                    elif element_d == "Expenses":
                        account_d = 'Salary Expense'
                    elif element_d == "Revenue":
                        account_d = revenue[1]
                elif event.key == pygame.K_3:
                    if toggle % 2 == 0:
                        element_d = "Shareholders' Equity"
                    elif element_d == 'Assets':
                        account_d = 'Notes Receivable'
                    elif element_d == 'Liabilities':
                        account_d = 'Interest Payable'
                    elif element_d == "Shareholders' Equity":
                        account_d = 'Dividends'
                    elif element_d == "Expenses":
                        account_d = expenses[2]
                elif event.key == pygame.K_4:
                    if toggle % 2 == 0:
                        element_d = "Expenses"
                    elif element_d == 'Assets':
                        account_d = 'Buildings'
                    elif element_d == 'Liabilities':
                        account_d = 'Taxes Payable'
                    elif element_d == "Shareholders' Equity":
                        account_d = 'Contributed Capital'
                    elif element_d == "Expenses":
                        account_d = expenses[3]
                elif event.key == pygame.K_5:
                    if toggle % 2 == 0:
                        element_d = "Revenue"
                    elif element_d == 'Assets':
                        account_d = 'Land'
                    elif element_d == 'Liabilities':
                        account_d = 'Bonds Payable'
                    elif element_d == "Expenses":
                        account_d = expenses[4]
                elif event.key == pygame.K_6:
                    if element_d == 'Assets':
                        account_d = 'Equipment'
                    elif element_d == 'Liabilities':
                        account_d = 'Deferred Revenue'
                    elif element_d == "Shareholders' Equity":
                        account_d = 'Rent Expense'
                elif event.key == pygame.K_7:
                    if element_d == 'Assets':
                        account_d = 'Prepaid Expenses'
                    elif element_d == 'Liabilities':
                        account_d = 'Loan Payable'
                    elif element_d == "Shareholders' Equity":
                        account_d = 'Service Revenue'
                elif event.key == pygame.K_8:
                    if element_d == 'Assets':
                        account_d = 'Supplies'
                    elif element_d == 'Liabilities':
                        account_d = 'Accrued Liabilities'
                    elif element_d == "Shareholders' Equity":
                        account_d = 'Sales Revenue'
                elif event.key == pygame.K_9:
                    if element_d == 'Assets':
                        account_d = 'Inventory'
                elif event.key == pygame.K_BACKSLASH:
                    if account_d != '':
                        debit = False
                        credit_stage()
                elif event.key == pygame.K_RIGHTBRACKET:
                    debit = False
                    element_d = ''
                    element_c = ''
                    account_d = ''
                    account_c = ''
                    value = ''
                    date = ''
                    description = ''
                    tab = 0
                    transScreen()

        gameDisplay.fill((255, 255, 255))
        button_ui(383, 25, 200, 50, text="Debit Account",text_size='medium')

        button_ui(120, 100, 115, 40, text="Assets", action='element_d_assets')
        button_ui(250, 100, 115, 40, text="Liabilities", action='element_d_liab')
        button_ui(380, 100, 115, 40, text="SE", action='element_d_SHE')
        button_ui(510, 100, 115, 40, text="Expenses", action='element_d_exp')
        button_ui(640, 100, 115, 40, text="Revenue", action='element_d_rev')

        button_ui(650, 600, 115, 40, text='Cancel', action='cancel')
        button_ui(800, 100, 115, 40, text='Done', action='debit_to_credit')

        if element_d == 'Assets':
            button_fill(193, 600, 250, 40, text='Other', action='other_asset', word=True)
            pygame.draw.line(gameDisplay, black, (120,100),(233,100),2)
            pygame.draw.line(gameDisplay, black, (120,138),(233,138),2)
            for i in range(len(asset)):
                button_ui(120 + (i * 130) % 390, 250 + 100 * (i // 3), 115, 40, text=asset[i],
                          action="account_d" + asset[i])
            if account_d != '':
                for j in range(len(asset)):
                    if account_d == asset[j]:
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 250 + 100 * (j // 3)),
                                         (118 + (j * 130) % 390 + 115, 250 + 100 * (j // 3)), 2)
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 248 + 100 * (j // 3) + 40),
                                         (118 + (j * 130) % 390 + 115, 248 + 100 * (j // 3) + 40), 2)
                        break
            if other_asset != []:
                for i in range(len(other_asset)):
                    button_ui(0, 160 + i * 45, 115, 40, text=other_asset[i],
                              action='account_d_' + other_asset[i])
                    for j in range(len(other_asset)):
                        if account_d == other_asset[j]:
                            pygame.draw.line(gameDisplay, black, (0, 160 + j * 45), (113, 160 + j * 45), 2)
                            pygame.draw.line(gameDisplay, black, (0, 198 + j * 45), (113, 198 + j * 45), 2)
                            break

        elif element_d == 'Liabilities':
            button_fill(193, 600, 250, 40, text='Other', action='other_liab', word=True)
            pygame.draw.line(gameDisplay, black, (250, 100), (363, 100), 2)
            pygame.draw.line(gameDisplay, black, (250, 138), (363, 138), 2)
            for i in range(len(liab)):
                button_ui(120 + (i * 130) % 390, 250 + 100 * (i // 3), 115, 40, text=liab[i],
                          action="account_d" + liab[i])
            if account_d != '':
                for j in range(len(liab)):
                    if account_d == liab[j]:
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 250 + 100 * (j // 3)),
                                         (118 + (j * 130) % 390 + 115, 250 + 100 * (j // 3)), 2)
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 248 + 100 * (j // 3) + 40),
                                         (118 + (j * 130) % 390 + 115, 248 + 100 * (j // 3) + 40), 2)
                        break
            if other_liab != []:
                for i in range(len(other_liab)):
                    button_ui(0, 160 + i * 45, 115, 40, text=other_liab[i],
                              action='account_d_' + other_liab[i])
                if account_d != '':
                    for j in range(len(other_liab)):
                        if account_d == other_liab[j]:
                            pygame.draw.line(gameDisplay, black, (0, 160 + j * 45), (113, 160 + j * 45), 2)
                            pygame.draw.line(gameDisplay, black, (0, 198 + j * 45), (113, 198 + j * 45), 2)
                            break

        elif element_d == "Shareholders' Equity":
            button_fill(193, 600, 250, 40, text='Other', action='other_SHE', word=True)
            pygame.draw.line(gameDisplay, black, (380, 100), (493, 100), 2)
            pygame.draw.line(gameDisplay, black, (380, 138), (493, 138), 2)
            for i in range(len(SHE)):
                button_ui(120 + (i * 130) % 390, 250 + 100 * (i // 3), 115, 40, text=SHE[i],
                          action="account_d" + SHE[i])
            if account_d != '':
                for j in range(len(SHE)):
                    if account_d == SHE[j]:
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 250 + 100 * (j // 3)),
                                         (118 + (j * 130) % 390 + 115, 250 + 100 * (j // 3)), 2)
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 248 + 100 * (j // 3) + 40),
                                         (118 + (j * 130) % 390 + 115, 248 + 100 * (j // 3) + 40), 2)
                        break
            if other_SHE != []:
                for i in range(len(other_SHE)):
                    button_ui(0, 160 + i * 45, 115, 40, text=other_SHE[i],
                              action='account_d_' + other_SHE[i])
                if account_d != '':
                    for j in range(len(other_SHE)):
                        if account_d == other_SHE[j]:
                            pygame.draw.line(gameDisplay, black, (0, 160 + j * 45), (113, 160 + j * 45), 2)
                            pygame.draw.line(gameDisplay, black, (0, 198 + j * 45), (113, 198 + j * 45), 2)
                            break

        elif element_d == 'Expenses':
            button_fill(193, 600, 250, 40, text='Other', action='other_expenses', word=True)
            pygame.draw.line(gameDisplay, black, (510, 100), (623, 100), 2)
            pygame.draw.line(gameDisplay, black, (510, 138), (623, 138), 2)
            for i in range(len(expenses)):
                button_ui(120 + (i * 130) % 390, 250 + 100 * (i//3), 115, 40, text=expenses[i], action = "account_d" + expenses[i])
            if account_d != '':
                for j in range(len(expenses)):
                    if account_d == expenses[j]:
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 250 + 100 * (j // 3)),
                                         (118 + (j * 130) % 390 + 115, 250 + 100 * (j // 3)), 2)
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 248 + 100 * (j // 3) + 40),
                                         (118 + (j * 130) % 390 + 115, 248 + 100 * (j // 3) + 40), 2)
                        break
            if other_expenses != []:
                for i in range(len(other_expenses)):
                    button_ui(0, 160 + i * 45, 115, 40, text=other_expenses[i], action='account_d_' + other_expenses[i])
                if account_d != '':
                    for j in range(len(other_expenses)):
                        if account_d == other_expenses[j]:
                            pygame.draw.line(gameDisplay, black, (0, 160 + j * 45), (113, 160 + j * 45), 2)
                            pygame.draw.line(gameDisplay, black, (0, 198 + j * 45), (113, 198 + j * 45), 2)
                            break

        elif element_d == 'Revenue':
            button_fill(193, 600, 250, 40, text='Other', action='other_revenue', word=True)
            pygame.draw.line(gameDisplay, black, (640, 100), (753, 100), 2)
            pygame.draw.line(gameDisplay, black, (640, 138), (753, 138), 2)
            for i in range(len(revenue)):
                button_ui(120 + (i * 130) % 390, 250 + 100 * (i // 3), 115, 40, text=revenue[i],
                          action="account_d" + revenue[i])
            if account_d != '':
                for j in range(len(revenue)):
                    if account_d == revenue[j]:
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 250 + 100 * (j // 3)),
                                         (118 + (j * 130) % 390 + 115, 250 + 100 * (j // 3)), 2)
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 248 + 100 * (j // 3) + 40),
                                         (118 + (j * 130) % 390 + 115, 248 + 100 * (j // 3) + 40), 2)
                        break
            if other_revenue != []:
                for i in range(len(other_revenue)):
                    button_ui(0, 160 + i * 45, 115, 40, text=other_revenue[i],
                              action='account_d_' + other_revenue[i])
                if account_d != '':
                    for j in range(len(other_revenue)):
                        if account_d == other_revenue[j]:
                            pygame.draw.line(gameDisplay, black, (0, 160 + j * 45), (113, 160 + j * 45), 2)
                            pygame.draw.line(gameDisplay, black, (0, 198 + j * 45), (113, 198 + j * 45), 2)
                            break

        if transaction_list != []:
            num = 0
            for i in range(len(transaction_list)):
                button_display(850,0 + num * 45,i)
                num += 1

        button_ui(1306, 0, 60, 25, text=chr(215), action="quit", text_size="large")
        button_ui(1246, 0, 60, 25, text=chr(8213), action="minimize")

        pygame.display.update()
        clock.tick(25)

def credit_stage():
    global element_c
    global account_c
    global toggle
    credit = True
    toggle = 0
    while credit:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                quit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    pygame.quit()
                    quit()
                elif event.key == pygame.K_SPACE:
                    toggle += 1
                    account_c = ''
                elif event.key == pygame.K_1:
                    if toggle % 2 == 0:
                        element_c = 'Assets'
                    elif element_c == 'Assets':
                        account_c = 'Cash'
                    elif element_c == 'Liabilities':
                        account_c = 'Account Payable'
                    elif element_c == "Shareholders' Equity":
                        account_c = 'Retained Earnings'
                    elif element_c == "Expenses":
                        account_c = 'Cost of Goods Sold'
                    elif element_c == "Revenue":
                        account_c = revenue[0]
                elif event.key == pygame.K_2:
                    if toggle % 2 == 0:
                        element_c = 'Liabilities'
                    elif element_c == 'Assets':
                        account_c = 'Account Receivable'
                    elif element_c == 'Liabilities':
                        account_c = 'Notes Payable'
                    elif element_c == "Shareholders' Equity":
                        account_c = 'Common Shares'
                    elif element_c == "Expenses":
                        account_c = expenses[1]
                    elif element_c == "Revenue":
                        account_c = revenue[1]
                elif event.key == pygame.K_3:
                    if toggle % 2 == 0:
                        element_c = "Shareholders' Equity"
                    elif element_c == 'Assets':
                        account_c = 'Notes Receivable'
                    elif element_c == 'Liabilities':
                        account_c = 'Interest Payable'
                    elif element_c == "Shareholders' Equity":
                        account_c = 'Dividends'
                    elif element_c == "Expenses":
                        account_c = expenses[2]
                elif event.key == pygame.K_4:
                    if toggle % 2 == 0:
                        element_c = 'Expenses'
                    elif element_c == 'Assets':
                        account_c = 'Buildings'
                    elif element_c == 'Liabilities':
                        account_c = 'Taxes Payable'
                    elif element_c == "Shareholders' Equity":
                        account_c = 'Contributed Capital'
                    elif element_c == "Expenses":
                        account_c = expenses[3]
                elif event.key == pygame.K_5:
                    if toggle % 2 == 0:
                        element_c = 'Revenue'
                    elif element_c == 'Assets':
                        account_c = 'Land'
                    elif element_c == 'Liabilities':
                        account_c = 'Bonds Payable'
                    elif element_c == "Expenses":
                        account_c = expenses[4]
                elif event.key == pygame.K_6:
                    if element_c == 'Assets':
                        account_c = 'Equipment'
                    elif element_c == 'Liabilities':
                        account_c = 'Deferred Revenue'
                    elif element_c == "Shareholders' Equity":
                        account_c = 'Rent Expense'
                elif event.key == pygame.K_7:
                    if element_c == 'Assets':
                        account_c = 'Prepaid Expenses'
                    elif element_c == 'Liabilities':
                        account_c = 'Loan Payable'
                    elif element_c == "Shareholders' Equity":
                        account_c = 'Service Revenue'
                elif event.key == pygame.K_8:
                    if element_c == 'Assets':
                        account_c = 'Supplies'
                    elif element_c == 'Liabilities':
                        account_c = 'Accrued Liabilities'
                    elif element_c == "Shareholders' Equity":
                        account_c = 'Sales Revenue'
                elif event.key == pygame.K_9:
                    if element_c == 'Assets':
                        account_c = 'Inventory'

                elif event.key == pygame.K_BACKSLASH:
                    if account_c != '':
                        credit = False
                        detailsScreen()
                elif event.key == pygame.K_RIGHTBRACKET:
                    credit = False
                    debit_stage()

        gameDisplay.fill((255, 255, 255))
        button_ui(383,25, 200, 50, text="Credit Account",text_size='medium')

        button_ui(120, 100, 115, 40, text="Assets", action='element_c_assets')
        button_ui(250, 100, 115, 40, text="Liabilities", action='element_c_liab')
        button_ui(380, 100, 115, 40, text="SE", action='element_c_SHE')
        button_ui(510, 100, 115, 40, text="Expenses", action='element_c_exp')
        button_ui(640, 100, 115, 40, text="Revenue", action='element_c_rev')

        button_ui(650,600,115,40,text='Back', action='back_from_credit')
        button_ui(800, 100, 115, 40, text='Done', action='credit_to_details')

        if element_c == 'Assets':
            button_fill(193, 600, 250, 40, text='Other', action='other_asset', word=True)
            pygame.draw.line(gameDisplay, black, (120,100),(233,100),2)
            pygame.draw.line(gameDisplay, black, (120,138),(233,138),2)
            for i in range(len(asset)):
                button_ui(120 + (i * 130) % 390, 250 + 100 * (i // 3), 115, 40, text=asset[i],
                          action="account_c" + asset[i])
            if account_c != '':
                for j in range(len(asset)):
                    if account_c == asset[j]:
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 250 + 100 * (j // 3)),
                                         (118 + (j * 130) % 390 + 115, 250 + 100 * (j // 3)), 2)
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 248 + 100 * (j // 3) + 40),
                                         (118 + (j * 130) % 390 + 115, 248 + 100 * (j // 3) + 40), 2)
                        break
            if other_asset != []:
                for i in range(len(other_asset)):
                    button_ui(0, 160 + i * 45, 115, 40, text=other_asset[i],
                              action='account_c_' + other_asset[i])
                if account_c != '':
                    for j in range(len(other_asset)):
                        if account_c == other_asset[j]:
                            pygame.draw.line(gameDisplay, black, (0, 160 + j * 45), (113, 160 + j * 45), 2)
                            pygame.draw.line(gameDisplay, black, (0, 198 + j * 45), (113, 198 + j * 45), 2)
                            break

        elif element_c == 'Liabilities':
            button_fill(193, 600, 250, 40, text='Other', action='other_liab', word=True)
            pygame.draw.line(gameDisplay, black, (250, 100), (363, 100), 2)
            pygame.draw.line(gameDisplay, black, (250, 138), (363, 138), 2)
            for i in range(len(liab)):
                button_ui(120 + (i * 130) % 390, 250 + 100 * (i // 3), 115, 40, text=liab[i],
                          action="account_c" + liab[i])
            if account_c != '':
                for j in range(len(liab)):
                    if account_c == liab[j]:
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 250 + 100 * (j // 3)),
                                         (118 + (j * 130) % 390 + 115, 250 + 100 * (j // 3)), 2)
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 248 + 100 * (j // 3) + 40),
                                         (118 + (j * 130) % 390 + 115, 248 + 100 * (j // 3) + 40), 2)
                        break
            if other_liab != []:
                for i in range(len(other_liab)):
                    button_ui(0, 160 + i * 45, 115, 40, text=other_liab[i],
                              action='account_c_' + other_liab[i])
                if account_c != '':
                    for j in range(len(other_liab)):
                        if account_c == other_liab[j]:
                            pygame.draw.line(gameDisplay, black, (0, 160 + j * 45), (113, 160 + j * 45), 2)
                            pygame.draw.line(gameDisplay, black, (0, 198 + j * 45), (113, 198 + j * 45), 2)
                            break

        elif element_c == "Shareholders' Equity":
            button_fill(193, 600, 250, 40, text='Other', action='other_SHE', word=True)
            pygame.draw.line(gameDisplay, black, (380, 100), (493, 100), 2)
            pygame.draw.line(gameDisplay, black, (380, 138), (493, 138), 2)
            for i in range(len(SHE)):
                button_ui(120 + (i * 130) % 390, 250 + 100 * (i // 3), 115, 40, text=SHE[i],
                          action="account_c" + SHE[i])
            if account_c != '':
                for j in range(len(SHE)):
                    if account_c == SHE[j]:
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 250 + 100 * (j // 3)),
                                         (118 + (j * 130) % 390 + 115, 250 + 100 * (j // 3)), 2)
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 248 + 100 * (j // 3) + 40),
                                         (118 + (j * 130) % 390 + 115, 248 + 100 * (j // 3) + 40), 2)
                        break
            if other_SHE != []:
                for i in range(len(other_SHE)):
                    button_ui(0, 160 + i * 45, 115, 40, text=other_SHE[i],
                              action='account_c_' + other_SHE[i])
                if account_c != '':
                    for j in range(len(other_SHE)):
                        if account_c == other_SHE[j]:
                            pygame.draw.line(gameDisplay, black, (0, 160 + j * 45), (113, 160 + j * 45), 2)
                            pygame.draw.line(gameDisplay, black, (0, 198 + j * 45), (113, 198 + j * 45), 2)
                            break

        elif element_c == 'Expenses':
            button_fill(193, 600, 250, 40, text='Other', action='other_expenses', word=True)
            pygame.draw.line(gameDisplay, black, (510, 100), (623, 100), 2)
            pygame.draw.line(gameDisplay, black, (510, 138), (623, 138), 2)
            for i in range(len(expenses)):
                button_ui(120 + (i * 130) % 390, 250 + 100 * (i//3), 115, 40, text=expenses[i], action = "account_c" + expenses[i])
            if account_c != '':
                for j in range(len(expenses)):
                    if account_c == expenses[j]:
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 250 + 100 * (j // 3)),
                                         (118 + (j * 130) % 390 + 115, 250 + 100 * (j // 3)), 2)
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 248 + 100 * (j // 3) + 40),
                                         (118 + (j * 130) % 390 + 115, 248 + 100 * (j // 3) + 40), 2)
                        break
            if other_expenses != []:
                for i in range(len(other_expenses)):
                    button_ui(0, 210 + i * 45, 115, 40, text=other_expenses[i], action='account_c_' + other_expenses[i])
                if account_c != '':
                    for j in range(len(other_expenses)):
                        if account_c == other_expenses[j]:
                            pygame.draw.line(gameDisplay, black, (0, 160 + j * 45), (113, 160 + j * 45), 2)
                            pygame.draw.line(gameDisplay, black, (0, 198 + j * 45), (113, 198 + j * 45), 2)
                            break

        elif element_c == 'Revenue':
            button_fill(193, 600, 250, 40, text='Other', action='other_revenue', word=True)
            pygame.draw.line(gameDisplay, black, (640, 100), (753, 100), 2)
            pygame.draw.line(gameDisplay, black, (640, 138), (753, 138), 2)
            for i in range(len(revenue)):
                button_ui(120 + (i * 130) % 390, 250 + 100 * (i // 3), 115, 40, text=revenue[i],
                          action="account_c" + revenue[i])
            if account_c != '':
                for j in range(len(revenue)):
                    if account_c == revenue[j]:
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 250 + 100 * (j // 3)),
                                         (118 + (j * 130) % 390 + 115, 250 + 100 * (j // 3)), 2)
                        pygame.draw.line(gameDisplay, black, (120 + (j * 130) % 390, 248 + 100 * (j // 3) + 40),
                                         (118 + (j * 130) % 390 + 115, 248 + 100 * (j // 3) + 40), 2)
                        break
            if other_revenue != []:
                for i in range(len(other_revenue)):
                    button_ui(0, 210 + i * 45, 115, 40, text=other_revenue[i],
                              action='account_c_' + other_revenue[i])
                if account_c != '':
                    for j in range(len(other_revenue)):
                        if account_c == other_revenue[j]:
                            pygame.draw.line(gameDisplay, black, (0, 160 + j * 45), (113, 160 + j * 45), 2)
                            pygame.draw.line(gameDisplay, black, (0, 198 + j * 45), (113, 198 + j * 45), 2)
                            break


        if len(transaction_list) != 0:
            num = 0
            for i in range(len(transaction_list)):
                button_display(850,0 + num * 45,i)
                num += 1

        button_ui(1306, 0, 60, 25, text=chr(215), action="quit", text_size="large")
        button_ui(1246, 0, 60, 25, text=chr(8213), action="minimize")

        pygame.display.update()
        clock.tick(25)

def button_fill(x, y, size_x, size_y, text=None, action=None, text_size='small',key=0, lock=1, lock_size = 6, word=False):
    global value
    global tab
    global date
    global description
    global comp_name
    global fs_date
    global dest_filename
    global save_location
    global pressed
    global typed
    global dep_ac_cost
    global dep_rs_cost
    global dep_time
    global dep_unit

    cursor = pygame.mouse.get_pos()
    click = pygame.mouse.get_pressed()
    if x < cursor[0] < x + size_x and y < cursor[1] < y + size_y or key % lock_size == lock:
        pygame.draw.line(gameDisplay, (196, 196, 196), (x, y), (x + size_x-2, y), 2)
        pygame.draw.line(gameDisplay, (196, 196, 196), (x, y + size_y - 2), (x + size_x-2, y + size_y - 2), 2)
        if click[0] == 1 and action != None or key % lock_size == lock:
            pressed = True
            typed = ''
            while pressed:
                cursor = pygame.mouse.get_pos()
                click = pygame.mouse.get_pressed()
                gameDisplay.fill((255, 255, 255))
                if click[0] == 1 and (x > cursor[0] or cursor[0] > x + size_x or y > cursor[1] or cursor[1] > y + size_y):
                    pressed = False
                    if typed == '':
                        typed = '0'
                button_ui(x, y, size_x, size_y, text=typed)
                pygame.draw.line(gameDisplay, black, (x, y), (x + size_x-2, y), 2)
                pygame.draw.line(gameDisplay, black, (x, y + size_y-2), (x + size_x-2, y + size_y-2), 2)
                keys = pygame.key.get_pressed()
                for event in pygame.event.get():
                    if keys[pygame.K_LSHIFT] != 0:
                        if keys[pygame.K_a] != 0:
                            typed += 'A'
                        elif keys[pygame.K_b] != 0:
                            typed += 'B'
                        elif keys[pygame.K_c] != 0:
                            typed += 'C'
                        elif keys[pygame.K_d] != 0:
                            typed += 'D'
                        elif keys[pygame.K_e] != 0:
                            typed += 'E'
                        elif keys[pygame.K_f] != 0:
                            typed += 'F'
                        elif keys[pygame.K_g] != 0:
                            typed += 'G'
                        elif keys[pygame.K_h] != 0:
                            typed += 'H'
                        elif keys[pygame.K_i] != 0:
                            typed += 'I'
                        elif keys[pygame.K_j] != 0:
                            typed += 'J'
                        elif keys[pygame.K_k] != 0:
                            typed += 'K'
                        elif keys[pygame.K_l] != 0:
                            typed += 'L'
                        elif keys[pygame.K_m] != 0:
                            typed += 'M'
                        elif keys[pygame.K_n] != 0:
                            typed += 'N'
                        elif keys[pygame.K_o] != 0:
                            typed += 'O'
                        elif keys[pygame.K_p] != 0:
                            typed += 'P'
                        elif keys[pygame.K_q] != 0:
                            typed += 'Q'
                        elif keys[pygame.K_r] != 0:
                            typed += 'R'
                        elif keys[pygame.K_s] != 0:
                            typed += 'S'
                        elif keys[pygame.K_t] != 0:
                            typed += 'T'
                        elif keys[pygame.K_u] != 0:
                            typed += 'U'
                        elif keys[pygame.K_v] != 0:
                            typed += 'V'
                        elif keys[pygame.K_w] != 0:
                            typed += 'W'
                        elif keys[pygame.K_x] != 0:
                            typed += 'X'
                        elif keys[pygame.K_y] != 0:
                            typed += 'Y'
                        elif keys[pygame.K_z] != 0:
                            typed += 'Z'
                    if event.type == pygame.QUIT:
                        pygame.quit()
                        quit()
                    if (event.type == pygame.KEYDOWN and keys[pygame.K_LSHIFT] == 0):
                        if event.key == pygame.K_0:
                            typed += '0'
                        elif event.key == pygame.K_1:
                            typed += '1'
                        elif event.key == pygame.K_2:
                            typed += '2'
                        elif event.key == pygame.K_3:
                            typed += '3'
                        elif event.key == pygame.K_4:
                            typed += '4'
                        elif event.key == pygame.K_5:
                            typed += '5'
                        elif event.key == pygame.K_6:
                            typed += '6'
                        elif event.key == pygame.K_7:
                            typed += '7'
                        elif event.key == pygame.K_8:
                            typed += '8'
                        elif event.key == pygame.K_9:
                            typed += '9'
                        elif event.key == pygame.K_SPACE:
                            typed += ' '
                        elif event.key == pygame.K_PERIOD:
                            typed += '.'
                        elif event.key == pygame.K_BACKSPACE:
                            typed = typed[:-1]
                        elif event.key == pygame.K_RETURN:
                            pressed = False
                            if typed == '':
                                typed = '0'
                        elif word == True:
                            if event.key == pygame.K_a:
                                typed += 'a'
                            elif event.key == pygame.K_b:
                                typed += 'b'
                            elif event.key == pygame.K_c:
                                typed += 'c'
                            elif event.key == pygame.K_d:
                                typed += 'd'
                            elif event.key == pygame.K_e:
                                typed += 'e'
                            elif event.key == pygame.K_f:
                                typed += 'f'
                            elif event.key == pygame.K_g:
                                typed += 'g'
                            elif event.key == pygame.K_h:
                                typed += 'h'
                            elif event.key == pygame.K_i:
                                typed += 'i'
                            elif event.key == pygame.K_j:
                                typed += 'j'
                            elif event.key == pygame.K_k:
                                typed += 'k'
                            elif event.key == pygame.K_l:
                                typed += 'l'
                            elif event.key == pygame.K_m:
                                typed += 'm'
                            elif event.key == pygame.K_n:
                                typed += 'n'
                            elif event.key == pygame.K_o:
                                typed += 'o'
                            elif event.key == pygame.K_p:
                                typed += 'p'
                            elif event.key == pygame.K_q:
                                typed += 'q'
                            elif event.key == pygame.K_r:
                                typed += 'r'
                            elif event.key == pygame.K_s:
                                typed += 's'
                            elif event.key == pygame.K_t:
                                typed += 't'
                            elif event.key == pygame.K_u:
                                typed += 'u'
                            elif event.key == pygame.K_v:
                                typed += 'v'
                            elif event.key == pygame.K_w:
                                typed += 'w'
                            elif event.key == pygame.K_x:
                                typed += 'x'
                            elif event.key == pygame.K_y:
                                typed += 'y'
                            elif event.key == pygame.K_z:
                                typed += 'z'

                button_ui(x+size_x,y+size_y/2-10,60,20,"Cancel",action='cancel_fill')
                pygame.display.update((x,y,size_x+60,size_y))
                clock.tick(25)
            if action == 'value':
                value = typed
                tab = 2
            elif action == 'date':
                date = typed
                tab = 4
            elif action == 'descrp':
                description = typed
                tab = 0
            elif action == 'comp_name':
                comp_name = typed
                tab = 2
            elif action == 'fs_date':
                fs_date = typed
                tab = 4
            elif action == 'other_asset':
                if (typed != '0' and not typed in all_asset):
                    other_asset.append(typed)
                    all_asset.append(typed)
            elif (action == 'other_liab' and not typed in all_liab):
                if typed != '0':
                    other_liab.append(typed)
                    all_liab.append(typed)
            elif (action == 'other_SHE' and not typed in all_SHE):
                if typed != '0':
                    other_SHE.append(typed)
                    all_SHE.append(typed)
            elif (action == 'other_expenses' and not typed in all_expenses):
                if typed != '0':
                    other_expenses.append(typed)
                    all_expenses.append(typed)
            elif (action == 'other_expenses' and not typed in all_revenue):
                if typed != '0':
                    other_revenue.append(typed)
                    all_revenue.append(typed)
            elif action == "file_name":
                dest_filename = typed
                tab = 6
            elif action == "set_accost":
                dep_ac_cost = float(typed)
                tab = 2
            elif action == "set_rscost":
                dep_rs_cost = float(typed)
                tab = 4
            elif action == "set_lttime":
                dep_time = int(typed)
                tab = 6
            elif action == "set_ltunit":
                dep_unit = float(typed)
                tab = 8


    else:
        button_ui(x, y, size_x, size_y)
    if text != None:
        text_to_button(text, black, x, y, size_x, size_y, text_size)

def button_ui(x, y, size_x, size_y, text = None, action=None, text_size = 'small',colour = black):
    global element_d
    global element_c
    global account_d
    global account_c
    global debit
    global credit
    global details
    global value
    global date
    global description
    global tab
    global selector
    global pressed
    global typed
    global trans
    global fin
    global acc_settings
    global bankrecon
    global analysis

    cursor = pygame.mouse.get_pos()
    click = pygame.mouse.get_pressed()
    if click[0] == 0:
        selector = True
    if action != None:
        if x < cursor[0] < x + size_x and y < cursor[1] < y + size_y:
            pygame.draw.line(gameDisplay,(196,196,196),(x,y),(x+size_x-2,y),2)
            pygame.draw.line(gameDisplay,(196,196,196),(x,y+size_y-2), (x+size_x-2,y+size_y-2), 2)
            if click[0] == 1 and selector and action != None:
                selector = False

                if 'element_d_' in action:
                    if 'assets' in action:
                        element_d = 'Assets'
                        account_d = ''
                    elif 'liab' in action:
                        element_d = 'Liabilities'
                        account_d = ''
                    elif 'SHE' in action:
                        element_d = "Shareholders' Equity"
                        account_d = ''
                    elif 'exp' in action:
                        element_d = 'Expenses'
                        account_d = ''
                    elif 'rev' in action:
                        element_d = 'Revenue'
                        account_d = ''

                elif 'account_d' in action:
                    if element_d == 'Assets':
                        for i in asset:
                            if i in action:
                                account_d = i
                                break
                        for j in other_asset:
                            if j in action:
                                account_d = j
                                break

                    elif element_d == 'Liabilities':
                        for i in liab:
                            if i in action:
                                account_d = i
                                break
                        for j in other_liab:
                            if j in action:
                                account_d = j
                                break

                    elif element_d == "Shareholders' Equity":
                        for i in SHE:
                            if i in action:
                                account_d = i
                                break
                        for j in other_SHE:
                            if j in action:
                                account_d = j
                                break

                    elif element_d == 'Expenses':
                        for i in expenses:
                            if i in action:
                                account_d = i
                                break
                        for j in other_expenses:
                            if j in action:
                                account_d = j
                                break
                    elif element_d == 'Revenue':
                        for i in revenue:
                            if i in action:
                                account_d = i
                                break
                        for j in other_revenue:
                            if j in action:
                                account_d = j
                                break

                elif 'element_c_' in action:
                    if 'assets' in action:
                        element_c = 'Assets'
                        account_c = ''
                    elif 'liab' in action:
                        element_c = 'Liabilities'
                        account_c = ''
                    elif 'SHE' in action:
                        element_c = "Shareholders' Equity"
                        account_c = ''
                    elif 'exp' in action:
                        element_c = 'Expenses'
                        account_c = ''
                    elif 'rev' in action:
                        element_c = 'Revenue'
                        account_c = ''

                elif 'account_c' in action:
                    if element_c == 'Assets':
                        for i in asset:
                            if i in action:
                                account_c = i
                                break
                        for j in other_asset:
                            if j in action:
                                account_c = j
                                break

                    elif element_c == 'Liabilities':
                        for i in liab:
                            if i in action:
                                account_c = i
                                break
                        for j in other_liab:
                            if j in action:
                                account_c = j
                                break

                    elif element_c == "Shareholders' Equity":
                        for i in SHE:
                            if i in action:
                                account_c = i
                                break
                        for j in other_SHE:
                            if j in action:
                                account_c = j
                                break

                    elif element_c == 'Expenses':
                        for i in expenses:
                            if i in action:
                                account_c = i
                                break
                        for j in other_expenses:
                            if j in action:
                                account_c = j
                                break
                    elif element_c == 'Revenue':
                        for i in revenue:
                            if i in action:
                                account_c = i
                                break
                        for j in other_revenue:
                            if j in action:
                                account_c = j
                                break

                elif action == 'debit_to_credit':
                    if account_d != '':
                        debit = False
                        credit_stage()
                elif action == 'credit_to_details':
                    if account_c != '' and account_d != '':
                        credit = False
                        detailsScreen()
                elif action == 'details_to_main':
                    if value != 0 and date != '':
                        details = False
                        transaction()
                        transScreen()
                elif action == 'cancel_fill':
                    pressed = False
                    typed = '0'
                elif action == 'new_trnsct':
                    element_d = ''
                    element_c = ''
                    account_d = ''
                    account_c = ''
                    value = ''
                    date = ''
                    description = ''
                    tab = 0
                    debit_stage()
                elif action == 'save':
                    saveScreen()
                elif action == 'confirm':
                    save()
                    transScreen()
                elif action == "quit":
                    pygame.quit()
                    quit()
                elif action == "minimize":
                    pygame.display.iconify()
                elif action == 'cancel':
                    debit = False
                    element_d = ''
                    element_c = ''
                    account_d = ''
                    account_c = ''
                    value = ''
                    date = ''
                    description = ''
                    tab = 0
                    transScreen()
                elif action == 'back_from_credit':
                    credit = False
                    debit_stage()
                elif action == 'back_from_details':
                    details = False
                    credit_stage()
                elif '_navi_from_' in action:
                    s = action.split('_navi_from_')
                    if s[1] == 'main':
                        trans = False
                    elif s[1] == 'fin':
                        fin = False
                    elif s[1] == 'acc_set':
                        acc_settings = False
                    elif s[1] == 'bankrecon':
                        bankrecon = False
                    elif s[1] == 'analysis':
                        analysis = False

                    if s[0] == 'main':
                        transScreen()
                    elif s[0] == 'fin':
                        fin_statement()
                    elif s[0] == 'acc_set':
                        account_settings()
                    elif s[0] == 'bankrecon':
                        bank_recon()
                    elif s[0] == 'analysis':
                        analysis_screen()




        if text != None:
            text_to_button(text,colour,x,y,size_x,size_y,text_size)

    else:
        pygame.draw.line(gameDisplay, (196, 196, 196), (x, y), (x + size_x-2, y), 2)
        pygame.draw.line(gameDisplay, (196, 196, 196), (x, y + size_y-2), (x + size_x-2, y + size_y-2), 2)
        if text != None:
            text_to_button(text,colour,x,y,size_x,size_y,text_size)

def button_confg(x, y, size_x, size_y, text = None, action=None, text_size = 'small',colour = black, bg_colour1 = None, bg_colour2 = None):
    global selector_confg
    global depreciation_account
    global depreciation
    global dep_method
    global dep_ac_cost
    global dep_rs_cost
    global dep_time
    global dep_unit

    cursor = pygame.mouse.get_pos()
    click = pygame.mouse.get_pressed()
    if click[0] == 0:
        selector_confg = True

    if action != None:
        if (bg_colour1 != None) and (bg_colour2 != None):
            if "set_dep_" in action:
                s = action.split("_dep_")
                if s[1] == dep_method:
                    pygame.draw.rect(gameDisplay,bg_colour2,(x,y,size_x, size_y))
                else:
                    pygame.draw.rect(gameDisplay, bg_colour1, (x, y, size_x, size_y))
            else:
                pygame.draw.rect(gameDisplay, bg_colour1, (x, y, size_x, size_y))

        if x < cursor[0] < x + size_x and y < cursor[1] < y + size_y:
            pygame.draw.line(gameDisplay,black,(x,y),(x+size_x-2,y),2)
            pygame.draw.line(gameDisplay,black,(x,y+size_y-2), (x+size_x-2,y+size_y-2), 2)
            if click[0] == 1 and selector and action != None:
                selector_confg = False
                if "accum_confg_" in action:
                    s = action.split("_confg_")
                    for i in all_asset:
                        if s[1] == i:
                            depreciation_account = i
                            gameDisplay.blit(transparency, (0,0))
                            pygame.display.update()
                            depreciationScreen()
                elif "set_dep_" in action:
                    if action == "set_dep_SL":
                        dep_method = "SL"
                    elif action == "set_dep_uop":
                        dep_method = "uop"
                    elif action == "set_dep_DDB":
                        dep_method = "DDB"
                elif action == "dep_to_acc":
                    depreciation = False
                    account_settings()
                elif action == "done_dep":
                    depreciation = False
                    if dep_method == "SL":
                        dep_expense = (dep_ac_cost - dep_rs_cost) / dep_time
                        output = []
                        for i in range(dep_time):
                            output.append(dep_expense)
                        dep_accounts[depreciation_account] = output
                    elif dep_method == "DDB":
                        sl_rate = 2 / dep_time
                        CV = dep_ac_cost
                        output = []
                        for j in range(dep_time-1):
                            dep_expense = CV * sl_rate
                            output.append(dep_expense)
                            CV -= dep_expense
                        output.append(CV - dep_rs_cost)
                        dep_accounts[depreciation_account] = output
                    else:
                        pass
                    dep_ac_cost = 0
                    dep_rs_cost = 0
                    dep_unit = 0
                    dep_time = 0
                    account_settings()




        if text != None:
            text_to_button(text, colour, x, y, size_x, size_y, text_size)

    else:
        pygame.draw.line(gameDisplay, (196, 196, 196), (x, y), (x + size_x-2, y), 2)
        pygame.draw.line(gameDisplay, (196, 196, 196), (x, y + size_y-2), (x + size_x-2, y + size_y-2), 2)
        if text != None:
            text_to_button(text, colour, x, y, size_x, size_y, text_size)

def button_display(x, y, index):
    credit_element = transaction_list[index][0]
    credit_account = transaction_list[index][1]
    debit_element = transaction_list[index][2]
    debit_account = transaction_list[index][3]
    amount = transaction_list[index][4]
    var_date = transaction_list[index][5]

    text_to_button(debit_element, black, x, y+2, 225, 15, 'tiny')
    text_to_button(debit_account, black, x, y+17, 225, 15, 'tiny')
    text_to_button(credit_element, black, x+225, y+2, 225, 15, 'tiny')
    text_to_button(credit_account, black, x+225, y+17, 225, 15, 'tiny')
    text_to_button(amount, black, x, y+32, 225, 15, 'tiny')
    text_to_button(var_date, black, x+225, y+32, 225, 15, 'tiny')
    text_to_button(str(index+1), black, x+430, y, 20, 20, 'tiny')
    pygame.draw.line(gameDisplay, black, (x, y), (x + 450, y), 2)
    pygame.draw.line(gameDisplay, black, (x, y+30), (x + 450, y+30), 2)
    pygame.draw.line(gameDisplay, black, (x, y+45), (x + 450, y+45), 2)
    pygame.draw.line(gameDisplay, black, (x, y), (x, y + 45), 2)
    pygame.draw.line(gameDisplay, black, (x+225, y), (x+225, y + 45), 2)
    pygame.draw.line(gameDisplay, black, (x+450, y), (x+450, y + 45), 2)
    pygame.draw.line(gameDisplay, black, (x + 430, y), (x+430, y + 20), 2)
    pygame.draw.line(gameDisplay, black, (x + 430, y+20), (x + 450, y + 20), 2)

def detailsScreen():
    global element_d
    global account_d
    global tab
    tab = 0
    details = True
    while details:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                quit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    pygame.quit()
                    quit()
                elif event.key == pygame.K_TAB:
                    tab += 1
                elif event.key == pygame.K_BACKSLASH:
                    if value != 0 and date != '':
                        details = False
                        transaction()
                        transScreen()
                elif event.key == pygame.K_RIGHTBRACKET:
                    details = False
                    credit_stage()

        gameDisplay.fill((255, 255, 255))
        button_ui(583,25,200,50,text="Details", text_size='medium')

        button_fill(300,200,250,40,text='Value: ' + value,action='value', key=tab, lock=1)
        button_fill(300, 250, 250, 40, text='Date: ' + date, action='date', key=tab, lock=3, word=True)
        button_fill(300, 300, 250, 40, text='Description: ' + description, action='descrp', key=tab, lock=5, word=True)

        button_ui(650,600,115,40,text='Back', action='back_from_details')
        button_ui(650, 100, 115, 40, text='Input', action='details_to_main')
        if len(transaction_list) != 0:
            num = 0
            for i in range(len(transaction_list)):
                button_display(850,0 + num * 45,i)
                num += 1

        button_ui(1306, 0, 60, 25, text=chr(215), action="quit", text_size="large")
        button_ui(1246, 0, 60, 25, text=chr(8213), action="minimize")

        pygame.display.update()
        clock.tick(25)

def transScreen():
    global element_d
    global element_c
    global account_d
    global account_c
    global details
    global value
    global date
    global tab
    global description
    global trans
    trans = True
    while trans:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                quit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    pygame.quit()
                    quit()
                elif event.key == pygame.K_BACKSLASH:
                    element_d = ''
                    element_c = ''
                    account_d = ''
                    account_c = ''
                    value = ''
                    date = ''
                    description = ''
                    tab = 0
                    debit_stage()
                elif event.key == pygame.K_INSERT:
                    saveScreen()

        gameDisplay.fill((255,255,255))
        button_ui(583,25,200,50,text="Transactions", text_size='medium')

        if len(transaction_list) != 0:
            num = 0
            if len(transaction_list) > 16:
                adjuster = len(transaction_list) - 16
            else:
                adjuster = 0
            for i in range(len(transaction_list)):
                button_display(850,0 + num * 45,i + adjuster)
                num += 1
        button_ui(350, 200, 115, 40, text='New Entry', action='new_trnsct')
        button_ui(350, 100, 115, 40, text='Save', action='save')
        transpanel()

        button_ui(1306,0,60,25,text=chr(215), action="quit",text_size="large")
        button_ui(1246, 0, 60, 25, text=chr(8213), action="minimize")

        pygame.display.update()
        clock.tick(25)

def transpanel():
    pygame.draw.rect(gameDisplay,(20,25,50),[0,0,200,768])
    button_ui(25,35,150,49,"Transactions",action=None,text_size='medium',colour=white)
    button_ui(25, 115, 150, 50, "Accounts Configuration", action="acc_set_navi_from_main", text_size='medium',
              colour=white)
    button_ui(25,195,150,50,"Statements",action="fin_navi_from_main", text_size='medium',colour=white)
    button_ui(25, 275, 150, 50, "Analysis", action="analysis_navi_from_main", text_size='medium',colour=white)
    button_ui(25,355,150,50, "Bank Reconciliation", action="bankrecon_navi_from_main", text_size="medium",colour=white)

def fin_statement():

    fin = True
    while fin:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                quit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    pygame.quit()
                    quit()

        gameDisplay.fill((255, 255, 255))
        button_ui(583, 25, 200, 50, text="Financial Statements", text_size='medium')
        finpanel()

        button_ui(1306, 0, 60, 25, text=chr(215), action="quit", text_size="large")
        button_ui(1246, 0, 60, 25, text=chr(8213), action="minimize")

        pygame.display.update()
        clock.tick(25)

def finpanel():
    pygame.draw.rect(gameDisplay, (20, 25, 50), [0, 0, 200, 768])
    button_ui(25, 35, 150, 49, "Transactions", action="main_navi_from_fin", text_size='medium', colour=white)
    button_ui(25, 115, 150, 50, "Accounts Configuration", action="acc_set_navi_from_fin", text_size='medium',
              colour=white)
    button_ui(25, 195, 150, 50, "Statements", action=None, text_size='medium', colour=white)
    button_ui(25, 275, 150, 50, "Analysis", action="analysis_navi_from_fin", text_size='medium', colour=white)
    button_ui(25, 355, 150, 50, "Bank Reconciliation", action="bankrecon_navi_from_fin", text_size="medium",
              colour=white)

def bank_recon():
    global bankrecon
    bankrecon = True
    while bankrecon:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                quit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    pygame.quit()
                    quit()

        gameDisplay.fill((255, 255, 255))
        button_ui(583, 25, 200, 50, text="Bank Reconciliation", text_size='medium')
        bankrecon_panel()

        button_ui(1306, 0, 60, 25, text=chr(215), action="quit", text_size="large")
        button_ui(1246, 0, 60, 25, text=chr(8213), action="minimize")

        pygame.display.update()
        clock.tick(25)

def bankrecon_panel():
    pygame.draw.rect(gameDisplay, (20, 25, 50), [0, 0, 200, 768])
    button_ui(25, 35, 150, 49, "Transactions", action="main_navi_from_bankrecon", text_size='medium', colour=white)
    button_ui(25, 115, 150, 50, "Accounts Configuration", action="acc_set_navi_from_bankrecon", text_size='medium',
              colour=white)
    button_ui(25, 195, 150, 50, "Statements", action="fin_navi_from_bankrecon", text_size='medium', colour=white)
    button_ui(25, 275, 150, 50, "Analysis", action="analysis_navi_from_bankrecon", text_size='medium', colour=white)
    button_ui(25, 355, 150, 50, "Bank Reconciliation", action=None, text_size="medium", colour=white)

def analysis_screen():
    global analysis
    analysis = True
    while analysis:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                quit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    pygame.quit()
                    quit()

        gameDisplay.fill((255, 255, 255))
        button_ui(583, 25, 200, 50, text="Analysis", text_size='medium')
        analysis_panel()

        button_ui(1306, 0, 60, 25, text=chr(215), action="quit", text_size="large")
        button_ui(1246, 0, 60, 25, text=chr(8213), action="minimize")

        pygame.display.update()
        clock.tick(25)

def analysis_panel():
    pygame.draw.rect(gameDisplay, (20, 25, 50), [0, 0, 200, 768])
    button_ui(25, 35, 150, 49, "Transactions", action="main_navi_from_analysis", text_size='medium', colour=white)
    button_ui(25, 115, 150, 50, "Accounts Configuration", action="acc_set_navi_from_analysisn", text_size='medium',
              colour=white)
    button_ui(25, 195, 150, 50, "Statements", action="fin_navi_from_analysis", text_size='medium', colour=white)
    button_ui(25, 275, 150, 50, "Analysis", action=None, text_size='medium', colour=white)
    button_ui(25, 355, 150, 50, "Bank Reconciliation", action="bankrecon_navi_from_analysis", text_size="medium", colour=white)

def saveScreen():
    global tab
    global comp_name
    global fs_date
    saveScreen = True
    tab = 0
    while saveScreen:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                quit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    pygame.quit()
                    quit()
                elif event.key == pygame.K_TAB:
                    tab += 1
                elif event.key == pygame.K_BACKSLASH:
                    save()
                    transScreen()

        gameDisplay.fill((255, 255, 255))

        button_fill(300, 200, 250, 40, text='Company Name: ' + comp_name, action='comp_name', key=tab, lock=1, word=True)
        button_fill(300, 300, 250, 40, text='Date: ' + fs_date, action='fs_date', key=tab, lock=3,word=True)
        button_fill(300, 400, 250, 40, text='File Name: ' + dest_filename, action='file_name', key=tab, lock=5,word=True)

        button_ui(350, 600, 115, 40, text='Confirm', action='confirm')

        button_ui(1306, 0, 60, 25, text=chr(215), action="quit", text_size="large")
        button_ui(1246, 0, 60, 25, text=chr(8213), action="minimize")

        pygame.display.update()
        clock.tick(25)

def account_settings():
    global acc_settings
    acc_settings = True

    others_acc_sum = 0
    for i in other_asset:
        others_acc_sum += 1
    for j in other_liab:
        others_acc_sum += 1
    for k in other_SHE:
        others_acc_sum += 1

    while acc_settings:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                quit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    pygame.quit()
                    quit()

        gameDisplay.fill((255, 255, 255))
        button_ui(583, 25, 200, 50, text="Accounts Configuration", text_size='medium')
        button_ui(1216,25,100,40,text='Back', action='main_navi_from_acc_set')

        button_ui(80,80,200,40,text="Assets")
        for i in range(len(asset)+len(other_asset)-1):
            if i < len(asset):
                button_ui(40,130 + i * 50,115,40,text=asset[i])
                button_confg(165, 137 + i * 50, 25, 25, text=None, action="accum_confg_" + asset[i-1], bg_colour1=(220,220,220), bg_colour2=(40,100,200))
                if accounts_set != set() and asset[i] in accounts_set:
                    pygame.draw.line(gameDisplay, green, (40, 130 + i * 50), (165, 130 + i * 50), 3)
                    pygame.draw.line(gameDisplay, green, (40, 170 + i * 50), (165, 170 + i * 50), 3)
            else:
                button_ui(40,130 + i * 50,115,40,text=other_asset[i-len(asset)])
        for j in range(len(liab)+len(other_liab)-1):
            if j < len(liab):
                button_ui(450, 130 + j * 50, 115, 40, text=liab[j])
                if accounts_set != set() and liab[j] in accounts_set:
                    pygame.draw.line(gameDisplay, green, (450, 130 + j * 50), (565, 130 + j * 50), 3)
                    pygame.draw.line(gameDisplay, green, (450, 170 + j * 50), (565, 170 + j * 50), 3)
            else:
                button_ui(450, 130 + j * 50, 115, 40, text=other_liab[j - len(liab)])
        for k in range(len(SHE)+len(other_SHE)-1):
            if k < len(SHE):
                button_ui(800, 130 + k * 50, 115, 40, text=SHE[k])
                if accounts_set != set() and SHE[k] in accounts_set:
                    pygame.draw.line(gameDisplay, green, (800, 130 + k * 50), (915, 130 + k * 50), 3)
                    pygame.draw.line(gameDisplay, green, (800, 170 + k * 50), (915, 170 + k * 50), 3)
            else:
                button_ui(800, 130 + k * 50, 115, 40, text=other_SHE[k - len(SHE)])

        button_ui(1306, 0, 60, 25, text=chr(215), action="quit", text_size="large")
        button_ui(1246, 0, 60, 25, text=chr(8213), action="minimize")

        pygame.display.update()
        clock.tick(25)

def acc_setting_panel():
    button_ui(25,35,150,49,"Transactions",action="main_navi_from_acc_set",text_size='medium',colour=white)
    button_ui(25,115,150,50,"Statements",action="fin_navi_from_acc_set", text_size='medium',colour=white)
    button_ui(25, 195, 150, 50, "Analysis", action="analysis", text_size='medium',colour=white)
    button_ui(25, 275, 150, 50, "Accounts Configuration", action=None, text_size='medium', colour=white)

def depreciationScreen():
    global tab

    tab = 0

    depreciation = True
    while depreciation:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                quit()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_ESCAPE:
                    pygame.quit()
                    quit()
                elif event.key == pygame.K_TAB:
                    tab += 1
                elif event.key == pygame.K_RIGHTBRACKET:
                    depreciation = False
                    account_settings()
        pygame.draw.rect(gameDisplay, (255,255,255), (383,234,600,300))

        button_confg(625,234,115,40,text=depreciation_account, text_size='small')
        button_confg(458,314,20,20,text=None, action='set_dep_SL',bg_colour1=(155,155,155), bg_colour2=(50,100,200))
        button_confg(483,314,115,40,text="Straight-Line")
        button_confg(623, 314, 20, 20, text=None, action='set_dep_uop', bg_colour1=(155, 155, 155),
                     bg_colour2=(50, 100, 200))
        button_confg(648, 314, 115, 40, text="UoP")
        button_confg(793, 314, 20, 20, text=None, action='set_dep_DDB', bg_colour1=(155, 155, 155),
                     bg_colour2=(50, 100, 200))
        button_confg(818, 314, 115, 40, text="DDB")

        button_fill(433,400,200,40,text="Acquisition Cost: " + str(dep_ac_cost) , action = "set_accost",key=tab, lock = 1, lock_size= 8)
        button_fill(433, 450, 200, 40, text="Residual Cost: " + str(dep_rs_cost), action="set_rscost", key=tab,
                    lock=3,lock_size= 8)
        button_fill(733, 400, 200, 40, text="Lifetime (months): " + str(dep_time), action="set_lttime", key=tab,
                    lock=5,lock_size= 8)
        button_fill(733, 450, 200, 40, text="Lifetime (units): " + str(dep_unit), action="set_ltunit", key=tab,
                    lock=7,lock_size= 8)

        button_confg(768, 494, 100, 40, text="Done", action="done_dep")
        button_confg(883,494,100,40,text="Back", action = "dep_to_acc")

        pygame.display.update((383,234,600,300))
        clock.tick(25)

def transaction():
    global line
    line += 1
    transaction_list.append([element_d, account_d, element_c, account_c, value, date,description,line])
    accounts_set.add(account_d)
    accounts_set.add(account_c)

def save():
    f = open(dest_filename + ".csv", 'wb')
    csv.register_dialect('colons', delimiter = ':')

    input_trnsct()

    make_ledger()
            
    ledger_balance_count()

    unadj_trial_balance()

    income_statement()

    statement_of_RE()

    balance_sheet()

    try:
        wb.save(filename=dest_filename + '.xlsx')
    except IOError:
        print("The file is currently open and cannot be edited.")

def input_trnsct():
    global wb
    global ws1
    global value_dict
    global sum_dict
    global accounts

    wb = Workbook()

    value_dict = {}
    sum_dict = {}
    accounts = []

    ws1 = wb.active
    ws1.title = 'Journal Entry'

    ws1['B1'] = "Journal Entry"
    ws1['B1'].alignment = Alignment(horizontal="center", vertical="center")
    ws1['B1'].font = Font(bold=True)

    ws1['A3'] = "Date"
    ws1['A3'].alignment = Alignment(horizontal="center", vertical="center")

    ws1['B3'] = "Accounts and Explanation"
    ws1['B3'].alignment = Alignment(horizontal="center", vertical="center")
    ws1.column_dimensions['B'].width = 35

    ws1['C3'] = "Debit"
    ws1['C3'].alignment = Alignment(horizontal="center", vertical="center")

    ws1['D3'] = "Credit"
    ws1['D3'].alignment = Alignment(horizontal="center", vertical="center")

    for i in range(len(transaction_list)):
        line_i = transaction_list[i][7]
        ws1['B' + str(line_i * 4)] = transaction_list[i][1]
        ws1['A' + str(line_i * 4)] = transaction_list[i][5]
        ws1['C' + str(line_i * 4)] = float(transaction_list[i][4])
        ws1['B' + str(line_i * 4 + 1)] = '     ' + transaction_list[i][3]
        ws1['D' + str(line_i * 4 + 1)] = float(transaction_list[i][4])
        ws1['A' + str(line_i * 4 + 1)] = '(' + str(line) + ')'
        ws1['B' + str(line_i * 4 + 2)] = transaction_list[i][6]
        ws1['B' + str(line_i * 4 + 2)].font = Font(italic=True)
        for j in range(3):
            ws1['C' + str(line_i * 4 + j)].number_format = '#,##0'
            ws1['D' + str(line_i * 4 + j)].number_format = '#,##0'

def make_ledger():
    global ws2
    ws2 = wb.create_sheet(title = 'Ledger')
    
    ws2.merge_cells('A1:D1')
    ws2['A1'] = "Ledger"
    ws2['A1'].font = Font(bold=True)
    ws2['A1'].alignment = Alignment(horizontal="center", vertical="center")

    bottom_border = Border(bottom=Side(style='thick'))
    left_border = Border(left=Side(style='thick'))
    right_border = Border(right=Side(style='thick'))
    
    for i in range(len(transaction_list)):
        cell_debit = transaction_list[i][1]
        cell_credit = transaction_list[i][3]
        cell_formula = "='Journal Entry'!C%i" % ((i + 1) * 4)
        cell_value = transaction_list[i][4]

        if not cell_debit in accounts:
            accounts.append(cell_debit)
            value_dict[cell_debit] = [[],[]]
            value_dict[cell_debit][0].append(float(cell_value))
            ws2[str(get_column_letter(len(accounts)*4 - 3)) + '4'] = '(' + str(i+1) + ')'
            ws2.merge_cells(str(get_column_letter(len(accounts)*4 - 2)) + '3:' + str(get_column_letter(len(accounts)*4 - 1)) + '3')
            ws2.column_dimensions[str(get_column_letter(len(accounts)*4 - 2))].width = 14
            ws2.column_dimensions[str(get_column_letter(len(accounts)*4 - 1))].width = 14
            ws2.column_dimensions[str(get_column_letter(len(accounts)*4 - 3))].width = 5
            ws2.column_dimensions[str(get_column_letter(len(accounts)*4))].width = 5
            ws2[str(get_column_letter(len(accounts)*4 - 2)) + '3'].alignment = Alignment(horizontal="center", vertical="center")
            ws2[str(get_column_letter(len(accounts)*4 - 2)) + '3'] = str(cell_debit)
            ws2[str(get_column_letter(len(accounts)*4 - 2)) + '4'] = cell_formula
            ws2[str(get_column_letter(len(accounts)*4 - 2)) + '4'].number_format = '#,##0'
            ws2[str(get_column_letter(len(accounts)*4 - 2)) + '3'].border = bottom_border
            ws2[str(get_column_letter(len(accounts)*4 - 1)) + '3'].border = bottom_border
            ws2[str(get_column_letter(len(accounts)*4 - 2)) + '4'].border = right_border
            
        else:
            value_dict[cell_debit][0].append(float(cell_value))
            column = accounts.index(cell_debit) * 4
            row = 4
            while not ws2[str(get_column_letter(column+2)) + str(row)].value == None:
                row += 1
            ws2[str(get_column_letter(column+1)) + str(row)] = '(' + str(i+1) + ')'
            ws2[str(get_column_letter(column+2)) + str(row)] = cell_formula
            ws2[str(get_column_letter(column+2)) + str(row)].number_format = '#,##0'
            ws2[str(get_column_letter(column+2)) + str(row)].border = right_border
            
        if not cell_credit in accounts:
            accounts.append(cell_credit)
            value_dict[cell_credit] = [[],[]]
            value_dict[cell_credit][1].append(float(cell_value))
            ws2[str(get_column_letter(len(accounts)*4)) + '4'] = '(' + str(i+1) + ')'
            ws2.merge_cells(str(get_column_letter(len(accounts)*4 - 2)) + '3:' + str(get_column_letter(len(accounts)*4-1)) + '3')
            ws2.column_dimensions[str(get_column_letter(len(accounts)*4 - 2))].width = 14
            ws2.column_dimensions[str(get_column_letter(len(accounts)*4 - 1))].width = 14
            ws2.column_dimensions[str(get_column_letter(len(accounts)*4 - 3))].width = 5
            ws2.column_dimensions[str(get_column_letter(len(accounts)*4))].width = 5
            ws2[str(get_column_letter(len(accounts)*4 - 2)) + '3'].alignment = Alignment(horizontal="center", vertical="center")
            ws2[str(get_column_letter(len(accounts)*4 - 2)) + '3'] = str(cell_credit)
            ws2[str(get_column_letter(len(accounts)*4-1)) + '4'] = cell_formula
            ws2[str(get_column_letter(len(accounts)*4-1)) + '4'].number_format = '#,##0'
            ws2[str(get_column_letter(len(accounts)*4 - 2)) + '3'].border = bottom_border
            ws2[str(get_column_letter(len(accounts)*4-1)) + '3'].border = bottom_border
            ws2[str(get_column_letter(len(accounts)*4-1)) + '4'].border = left_border
            
        else:
            value_dict[cell_credit][1].append(float(cell_value))
            column = accounts.index(cell_credit) * 4
            row = 4
            while not ws2[str(get_column_letter(column+3)) + str(row)].value == None:
                row += 1
            ws2[str(get_column_letter(column+4)) + str(row)] = '(' + str(i+1) + ')'
            ws2[str(get_column_letter(column+3)) + str(row)] = cell_formula
            ws2[str(get_column_letter(column+3)) + str(row)].number_format = '#,##0'
            ws2[str(get_column_letter(column+3)) + str(row)].border = left_border

def ledger_balance_count():
    for i in range(1, len(accounts)+1):
        debit_count = 0
        credit_count = 0

        debit_sum = 0
        credit_sum = 0
        
        column = i * 4 - 2
        row_d = 3
        row_c = 3
        while not ws2[str(get_column_letter(column)) + str(row_d+1)].value == None:
            row_d += 1
            debit_count += 1
        while not ws2[str(get_column_letter(column+1)) + str(row_c+1)].value == None:
            row_c += 1
            credit_count += 1
        cell_name = ws2[str(get_column_letter(column)) + '3'].value
        if len(value_dict[cell_name][0]) > 0:
            for i in value_dict[cell_name][0]:
                debit_sum += i
        if len(value_dict[cell_name][1]) > 0:
            for j in value_dict[cell_name][1]:
                credit_sum += j
        checker = debit_sum - credit_sum
        if checker >= 0:
            ws2[str(get_column_letter(column)) + str(max(row_d+1,row_c+1))]='=SUM(%s4:%s%i)-SUM(%s4:%s%i)'%(get_column_letter(column),get_column_letter(column),max(row_d,row_c),get_column_letter(column+1),get_column_letter(column+1),max(row_d,row_c))
            ws2[str(get_column_letter(column-1)) + str(max(row_d+1,row_c+1))] = "Bal."
            ws2[str(get_column_letter(column)) + str(max(row_d+1,row_c+1))].font =Font(underline='doubleAccounting')
            sum_dict[cell_name] = [0,"='Ledger'!%s%d" % (get_column_letter(column), max(row_d+1,row_c+1))]
        else:
            ws2[str(get_column_letter(column+1)) + str(max(row_d+1,row_c+1))]='=SUM(%s4:%s%i)-SUM(%s4:%s%i)'%(get_column_letter(column+1),get_column_letter(column+1),max(row_d,row_c),get_column_letter(column),get_column_letter(column),max(row_d,row_c))
            ws2[str(get_column_letter(column+2)) + str(max(row_d+1,row_c+1))] = "Bal."
            ws2[str(get_column_letter(column+1)) + str(max(row_d+1,row_c+1))].font =Font(underline='doubleAccounting')
            sum_dict[cell_name] = [1,"='Ledger'!%s%d" % (get_column_letter(column+1), max(row_d+1,row_c+1))]

def unadj_trial_balance():
    ws3 = wb.create_sheet(title = 'Trial Balance')
    print("Accounts: " + str(accounts))
    print("sum_dict: " + str(sum_dict))
    print("value_dict" + str(value_dict))
    print("transaction_list: " + str(transaction_list))
    ws3.merge_cells('A1:C1')
    ws3['A1'] = comp_name
    ws3['A1'].font = Font(bold=True)
    ws3['A1'].alignment = Alignment(horizontal="center", vertical="center")

    ws3.merge_cells('A2:C2')
    ws3['A2'] = "Trial Balance"
    ws3['A2'].font = Font(bold=True)
    ws3['A2'].alignment = Alignment(horizontal="center", vertical="center")

    ws3.merge_cells('A3:C3')
    ws3['A3'] = fs_date
    ws3['A3'].font = Font(bold=True)
    ws3['A3'].alignment = Alignment(horizontal="center", vertical="center")

    ws3.column_dimensions['A'].width = 40
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 20

    ws3.merge_cells('A5:A6')
    ws3['A5'] = "Account Title"
    ws3['A5'].font = Font(bold=True)
    ws3['A5'].alignment = Alignment(horizontal="center", vertical="center")
    ws3.merge_cells('B5:C5')
    ws3['B5'] = "Balance"
    ws3['B5'].font = Font(bold=True)
    ws3['B5'].alignment = Alignment(horizontal="center", vertical="center")
    ws3['B6'] = "Debit"
    ws3['B6'].font = Font(bold=True)
    ws3['B6'].alignment = Alignment(horizontal="center", vertical="center")
    ws3['C6'] = "Credit"
    ws3['C6'].font = Font(bold=True)
    ws3['C6'].alignment = Alignment(horizontal="center", vertical="center")
    
    for count in range(3):
        if count == 0:
            ws3['A7'] = "Assets"
            ws3['A7'].font = Font(bold=True)
            ws3['A7'].alignment = Alignment(horizontal="center", vertical="center")
            for i in accounts:
                if i in all_asset:
                    row = find_row('A',8,ws3)
                    ws3['A' + str(row)] = i
                    if sum_dict[i][0] == 0:
                        ws3['B' + str(row)] = sum_dict[i][1]
                        ws3['B' + str(row)].number_format = '#,##0'
                    else:
                        ws3['C' + str(row)] = sum_dict[i][1]
                        ws3['C' + str(row)].number_format = '#,##0'
        elif count == 1:
            global starting_row
            starting_row = find_row('A',8,ws3)
            ws3['A' + str(starting_row)] = "Liabilities"
            ws3['A' + str(starting_row)].font = Font(bold=True)
            ws3['A' + str(starting_row)].alignment = Alignment(horizontal="center", vertical="center")
            for i in accounts:
                if i in all_liab:
                    row = find_row('A',starting_row,ws3)
                    ws3['A' + str(row)] = i
                    if sum_dict[i][0] == 0:
                        ws3['B' + str(row)] = sum_dict[i][1]
                        ws3['B' + str(row)].number_format = '#,##0'
                    else:
                        ws3['C' + str(row)] = sum_dict[i][1]
                        ws3['C' + str(row)].number_format = '#,##0'
       
        elif count == 2:
            starting_row = find_row('A',8,ws3)
            ws3['A' + str(starting_row)] = "Shareholders' Equity"
            ws3['A' + str(starting_row)].font = Font(bold=True)
            ws3['A' + str(starting_row)].alignment = Alignment(horizontal="center", vertical="center")
            for i in accounts:
                if i in all_SHE:
                    row = find_row('A',starting_row,ws3)
                    ws3['A' + str(row)] = i
                    if sum_dict[i][0] == 0:
                        ws3['B' + str(row)] = sum_dict[i][1]
                        ws3['B' + str(row)].number_format = '#,##0'
                    else:
                        ws3['C' + str(row)] = sum_dict[i][1]
                        ws3['C' + str(row)].number_format = '#,##0'
            for j in accounts:
                if j in all_revenue:
                    row = find_row('A',starting_row,ws3)
                    ws3['A' + str(row)] = j
                    if sum_dict[j][0] == 0:
                        ws3['B' + str(row)] = sum_dict[j][1]
                        ws3['B' + str(row)].number_format = '#,##0'
                    else:
                        ws3['C' + str(row)] = sum_dict[j][1]
                        ws3['C' + str(row)].number_format = '#,##0'
            for k in accounts:
                if k in all_expenses:
                    row = find_row('A',starting_row,ws3)
                    #row = starting_row
                    #while not ws3['A' + str(row)].value == None:
                    #    row += 1
                    ws3['A' + str(row)] = k
                    if sum_dict[k][0] == 0:
                        ws3['B' + str(row)] = sum_dict[k][1]
                        ws3['B' + str(row)].number_format = '#,##0'
                    else:
                        ws3['C' + str(row)] = sum_dict[k][1]
                        ws3['C' + str(row)].number_format = '#,##0'

            starting_row = find_row('A', starting_row, ws3)
            ws3['A' + str(starting_row)] = "Total"
            ws3['A' + str(starting_row)].font = Font(bold=True)
            ws3['B' + str(starting_row)] = "=SUM(B8:B%d)" % (starting_row-1)
            ws3['B' + str(starting_row)].font = Font(bold=True, underline='doubleAccounting')
            ws3['B' + str(starting_row)].number_format = '#,##0'
            ws3['C' + str(starting_row)] = "=SUM(C8:C%d)" % (starting_row-1)
            ws3['C' + str(starting_row)].font = Font(bold=True, underline='doubleAccounting')
            ws3['C' + str(starting_row)].number_format = '#,##0'

def income_statement():
    global net_income

    ws4 = wb.create_sheet(title='Income Statement')
    ws4.merge_cells('A1:C1')
    ws4.merge_cells('A2:C2')
    ws4.merge_cells('A3:C3')
    ws4['A1'] = comp_name
    ws4['A1'].font = Font(bold=True)
    ws4['A1'].alignment = Alignment(horizontal="center", vertical="center")

    ws4['A2'] = "Income Statement"
    ws4['A2'].font = Font(bold=True)
    ws4['A2'].alignment = Alignment(horizontal="center", vertical="center")

    ws4['A3'] = 'Month Ended ' + fs_date
    ws4['A3'].font = Font(bold=True)
    ws4['A3'].alignment = Alignment(horizontal="center", vertical="center")

    ws4.column_dimensions['A'].width = 40
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 20

    ws4['A5'] = "Revenues"
    ws4['A5'].font = Font(bold=True)


    gross_revenue = 0
    for i in accounts:
        if i in all_revenue:
            row_1 = find_row('A',6,ws4)
            ws4['A' + str(row_1)] = '   ' + i
            ws4['C' + str(row_1)] = sum_dict[i][1]
            ws4['C' + str(row_1)].number_format = '#,##0'
            for j in value_dict[i][0]:
                gross_revenue -= j
            for j in value_dict[i][1]:
                gross_revenue += j

    row_2 = find_row('A', 5, ws4)
    ws4['A' + str(row_2)] = "Expenses"
    ws4['A' + str(row_2)].font = Font(bold=True)

    exp_output = 0
    for j in accounts:
        if j in all_expenses:
            row_3 = find_row('A',row_2+1,ws4)
            ws4['A' + str(row_3)] = '   ' + j
            ws4['B' + str(row_3)] = sum_dict[j][1]
            ws4['B' + str(row_3)].number_format = '#,##0'
            for k in value_dict[j][0]:
                exp_output += k
            for k in value_dict[j][1]:
                exp_output -= k

    #print(str(row_2 + 1), str(find_row('B', row_2 + 1, ws4)-1))

    row_4 = find_row('A', row_2, ws4)
    if find_row('B',row_2+1,ws4) != row_2+1:
        total_expenses = ['=SUM(B%i:B%i)' % (row_2+1,find_row('B',row_2+1,ws4)-1)]
        total_expenses.append(exp_output)
    else:
        total_expenses = [0,0]

    ws4['A' + str(row_4)] = "   " + "Total Expenses"
    ws4['C' + str(row_4)] = total_expenses[0]
    ws4['C' + str(row_4)].number_format = '#,##0'
    ws4['C' + str(row_4)].font = Font(underline='singleAccounting')

    if row_2 - 1 == 5:
        net_income = ["=0 - C%i" % (row_4)]
    else:
        net_income = ["=SUM(C5:C%i) - C%i" % (row_2 - 1, row_4)]
    net_income.append(gross_revenue - total_expenses[1])
    ws4['A' + str(row_4 + 1)] = "Net Income"
    ws4['A' + str(row_4 + 1)].font = Font(bold=True)
    ws4['C' + str(row_4 + 1)] = net_income[0]
    ws4['C' + str(row_4 + 1)].font = Font(bold=True, underline='doubleAccounting')
    ws4['C' + str(row_4 + 1)].number_format = '#,##0'
    net_income.append("='Income Statement'!C" + str(row_4 + 1))
    #print(net_income[1], net_income[2])

def statement_of_RE():
    ws5 = wb.create_sheet(title='Statement of Retained Earnings')
    ws5.merge_cells('A1:B1')
    ws5.merge_cells('A2:B2')
    ws5.merge_cells('A3:B3')
    ws5['A1'] = comp_name
    ws5['A1'].font = Font(bold=True)
    ws5['A1'].alignment = Alignment(horizontal="center", vertical="center")

    ws5['A2'] = "Statement of Retained Earnings"
    ws5['A2'].font = Font(bold=True)
    ws5['A2'].alignment = Alignment(horizontal="center", vertical="center")

    ws5['A3'] = 'Month Ended ' + fs_date
    ws5['A3'].font = Font(bold=True)
    ws5['A3'].alignment = Alignment(horizontal="center", vertical="center")

    ws5.column_dimensions['A'].width = 40
    ws5.column_dimensions['B'].width = 20

    ws5['A5'] = "Retained Earnings, "
    ws5['A5'].font = Font(bold=True)
    if "Retained Earnings" in sum_dict:
        ws5['B5'] = sum_dict["Retained Earnings"][1]
        ws5['B5'].number_format = '#,##0'
    else:
        ws5['B5'] = 0
        ws5['B5'].number_format = '#,##0'

    ws5['A6'] = "Add: Net Income"
    ws5['B6'] = net_income[2]
    ws5['B6'].number_format = '#,##0'

    ws5['A7'] = "Less: Dividends Declared"
    if "Dividends" in sum_dict:
        ws5['B7'] = sum_dict["Dividends"][1]
        ws5['B7'].number_format = '#,##0'
    else:
        ws5['B7'] = 0
        ws5['B7'].number_format = ' #,##0'

    ws5['A8'] = "Retained Earnings, "
    ws5['B8'] = '=SUM(B5,B6) - B7'
    ws5['B8'].font = Font(bold=True, underline='doubleAccounting')
    ws5['B8'].number_format = '#,##0'

    if "Retained Earnings" in sum_dict:
        sum_dict["Retained Earnings"][1] = "='Statement of Retained Earnings'!B8"

def balance_sheet():
    upper_border = Border(upper=Side(style='thick'))

    ws6 = wb.create_sheet(title='Balance Sheet')
    ws6.merge_cells('A1:C1')
    ws6.merge_cells('A2:C2')
    ws6.merge_cells('A3:C3')
    ws6['A1'] = comp_name
    ws6['A1'].font = Font(bold=True)
    ws6['A1'].alignment = Alignment(horizontal="center", vertical="center")

    ws6['A2'] = "Balance Sheet"
    ws6['A2'].font = Font(bold=True)
    ws6['A2'].alignment = Alignment(horizontal="center", vertical="center")

    ws6['A3'] = fs_date
    ws6['A3'].font = Font(bold=True)
    ws6['A3'].alignment = Alignment(horizontal="center", vertical="center")

    ws6.column_dimensions['A'].width = 40
    ws6.column_dimensions['B'].width = 20
    ws6.column_dimensions['C'].width = 20

    ws6.merge_cells('A5:C5')
    ws6['A5'] = "Assets"
    ws6['A5'].font = Font(bold=True)
    ws6['A5'].alignment = Alignment(horizontal="center", vertical="center")

    for i in accounts:
        if i in all_asset:
            starting_row = 6
            row = find_row('A', starting_row, ws6)
            ws6['A' + str(row)] = i
            ws6['C' + str(row)] = sum_dict[i][1]
            ws6['C' + str(row)].number_format = '#,##0'

    ws6['A%i' % (row + 1)] = "Total Assets:"
    ws6['A%i' % (row + 1)].font = Font(bold=True)
    ws6['A%i' % (row + 1)].border = upper_border
    ws6['C%i' % (row + 1)] = "=SUM(C%i:C%i)" % (starting_row, row)
    ws6['C%i' % (row + 1)].number_format = '#,##0'

    ws6.merge_cells('A%i:C%i' % (row + 3, row + 3))
    ws6['A%i' % (row + 3)] = "Liabilities"
    ws6['A%i' % (row + 3)].font = Font(bold=True)
    ws6['A%i' % (row + 3)].alignment = Alignment(horizontal="center", vertical="center")

    starting_row = row + 4

    for j in accounts:
        if j in all_liab:
            row = find_row('A', starting_row, ws6)
            ws6['A' + str(row)] = j
            ws6['C' + str(row)] = sum_dict[j][1]
            ws6['C' + str(row)].number_format = '#,##0'

    ws6['A%i' % (row + 1)] = "Total Liabilities:"
    ws6['A%i' % (row + 1)].font = Font(bold=True)
    ws6['A%i' % (row + 1)].border = upper_border
    ws6['C%i' % (row + 1)] = "=SUM(C%i:C%i)" % (starting_row, row)
    ws6['C%i' % (row + 1)].number_format = '#,##0'

    ws6.merge_cells('A%i:C%i' % (row + 3, row + 3))
    ws6['A%i' % (row + 3)] = "Shareholders' Equity"
    ws6['A%i' % (row + 3)].font = Font(bold=True)
    ws6['A%i' % (row + 3)].alignment = Alignment(horizontal="center", vertical="center")

    starting_row = row + 4

    for k in accounts:
        if k in all_SHE:
            row = find_row('A', starting_row, ws6)
            ws6['A' + str(row)] = k
            ws6['C' + str(row)] = sum_dict[k][1]
            ws6['C' + str(row)].number_format = '#,##0'

    ws6['A%i' % (row + 1)] = "Total Shareholders' Equity:"
    ws6['A%i' % (row + 1)].font = Font(bold=True)
    ws6['A%i' % (row + 1)].border = upper_border
    ws6['C%i' % (row + 1)] = "=SUM(C%i:C%i)" % (starting_row, row)
    ws6['C%i' % (row + 1)].number_format = '#,##0'

debit_stage()




